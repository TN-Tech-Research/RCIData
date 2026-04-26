#!/usr/bin/env python3
"""
update_from_excel.py  —  Refresh index.html stats from Satisfaction.xlsx

Usage:
  python3 update_from_excel.py                      # uses Satisfaction.xlsx in same dir
  python3 update_from_excel.py path/to/file.xlsx    # use a different file

To auto-download before updating, set FORMS_EXCEL_URL in the environment:
  FORMS_EXCEL_URL="https://..." python3 update_from_excel.py
"""

import sys
import os
import re
import math
from collections import Counter, defaultdict

# ── Optional: download from Microsoft Forms / SharePoint ──────────────────────
FORMS_EXCEL_URL = os.environ.get("FORMS_EXCEL_URL", "")

def download_excel(url: str, dest: str) -> None:
    import urllib.request
    print(f"Downloading Excel from {url} ...")
    urllib.request.urlretrieve(url, dest)
    print(f"Saved to {dest}")

# ── Load workbook ─────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, "Satisfaction.xlsx")
HTML_PATH  = os.path.join(SCRIPT_DIR, "index.html")

if FORMS_EXCEL_URL:
    download_excel(FORMS_EXCEL_URL, EXCEL_PATH)

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
data = rows[1:]  # skip header row
N    = len(data)

# ── Column indices (0-based) ──────────────────────────────────────────────────
C_OVERALL  = 5
C_COMM     = 6
C_FACULTY  = 7
C_SP       = 8   # SharePoint
C_REG      = 9   # Editable registration form
C_MAP      = 10  # Table map
C_JUDGE    = 11  # Digital judging
C_FEEDBACK = 12  # Judge feedback
C_EMAIL    = 13  # Automated email
C_IMPROVE  = 14  # Improvements multi-select
C_HAD_ISSUE = 16
C_ISSUE_TYPE = 17
C_RESOLVED   = 18

# ── Helper: numeric ratings ───────────────────────────────────────────────────
def ratings(col):
    return [r[col] for r in data if isinstance(r[col], (int, float))]

def mean(lst):
    return sum(lst) / len(lst) if lst else 0

def pct_4_5(lst):
    return round(sum(1 for x in lst if x >= 4) / len(lst) * 100) if lst else 0

def dist(lst):
    c = Counter(lst)
    total = len(lst)
    result = {}
    for star in range(5, 0, -1):
        cnt = c.get(star, 0)
        result[star] = {"cnt": cnt, "pct": round(cnt / total * 100)}
    return result

# ── Slide 1 stats ─────────────────────────────────────────────────────────────
overall  = ratings(C_OVERALL)
comm_r   = ratings(C_COMM)
faculty  = ratings(C_FACULTY)

avg_all  = round((mean(overall) + mean(comm_r) + mean(faculty)) / 3, 2)
pct_ov   = pct_4_5(overall)
pct_no_issues = round(sum(1 for r in data if r[C_HAD_ISSUE] == 'No') / N * 100)

# ── Slide 2 stats ─────────────────────────────────────────────────────────────
ov_dist  = dist(overall)
cm_dist  = dist(comm_r)
fa_dist  = dist(faculty)
pct_45_comm = pct_4_5(comm_r)
pct_45_fac  = pct_4_5(faculty)
fa_n    = len(faculty)
fa5     = fa_dist[5]["cnt"]

# ── Slide 3: digital tools ────────────────────────────────────────────────────
TOOL_COLS = [C_SP, C_REG, C_MAP, C_JUDGE, C_FEEDBACK, C_EMAIL]
TOOL_NAMES = [
    "SharePoint event page",
    "Editable registration form",
    "Interactive table map",
    "Digital judging",
    "Option to receive judge feedback",
    "Automated email reminders",
]

def tool_stats(col):
    vals = [r[col] for r in data if r[col] is not None]
    total = len(vals)
    c = Counter(vals)
    vh  = c.get("Very helpful", 0)
    sh  = c.get("Somewhat helpful", 0)
    neu = c.get("Neither helpful nor problematic", 0)
    sp  = c.get("Somewhat problematic", 0)
    vp  = c.get("Very problematic", 0)
    return {
        "vh":  round(vh  / total * 100),
        "sh":  round(sh  / total * 100),
        "neu": round(neu / total * 100),
        "prob": round((sp + vp) / total * 100),
        "pos": round((vh + sh) / total * 100),
        "n": total,
    }

tools = [tool_stats(col) for col in TOOL_COLS]

# Identify top-2 by positive pct for gold highlights
sorted_by_pos = sorted(range(len(tools)), key=lambda i: tools[i]["pos"], reverse=True)
gold_tools = set(sorted_by_pos[:2])

# ── Slide 5: improvements ─────────────────────────────────────────────────────
IMPROVE_KEYS = {
    "Calendar invites for event times and locations": "Calendar invites for times &amp; locations",
    "Centralized portal for print requests and print previews.": "Centralized print portal",
    "Modernized poster templates on multiple platforms, such as Canva.": "Modernized templates (Canva)",
    "Collaborations with local industries or businesses": "Industry collaborations",
    "Use of Microsoft Teams for event information": "Teams for event information",
    "Faculty involvement in selection of judges": "Faculty involvement in judging",
    "Additional time for networking or input from faculty": "Additional networking time",
    "Addition of a Presentation category based on verbal communication": "Presentation/verbal category",
    "Additional digital tools": "Additional digital tools",
    "Addition of an interdepartmental robotics competition": "Interdepartmental robotics competition",
}

imp_counts = defaultdict(int)
imp_resp   = set()
for r in data:
    if r[C_IMPROVE]:
        items = [x.strip() for x in r[C_IMPROVE].split(";") if x.strip()]
        if items:
            imp_resp.add(r[0])
        for item in items:
            for key, label in IMPROVE_KEYS.items():
                if key in item:
                    imp_counts[label] += 1

n_imp = len(imp_resp)
improvements = sorted(
    [(label, cnt) for label, cnt in imp_counts.items()],
    key=lambda x: -x[1]
)

# ── Slide 6: issues ───────────────────────────────────────────────────────────
had_issues = [r for r in data if r[C_HAD_ISSUE] == "Yes"]
n_issues   = len(had_issues)

def is_printing(t):
    return t and ("rinting" in t)

def is_not_aware(t):
    return t and ("weren't aware" in t or "did not receive the email" in t)

def is_academic(t):
    return t and "academic obligations" in t

def is_ceremony(t):
    return t and any(k in t.lower() for k in ("ceremony", "award", "name")) and not is_printing(t) and not is_not_aware(t) and not is_academic(t)

def categorize(t):
    if is_printing(t):   return "Printing defect"
    if is_not_aware(t):  return "Missed event — not aware"
    if is_academic(t):   return "Missed event — academic conflict"
    if is_ceremony(t):   return "Award ceremony error"
    return "Other / unspecified"

issue_totals   = defaultdict(int)
issue_resolved = defaultdict(int)
ISSUE_ORDER    = ["Printing defect", "Missed event — not aware",
                  "Missed event — academic conflict", "Award ceremony error",
                  "Other / unspecified"]

for r in had_issues:
    cat = categorize(r[C_ISSUE_TYPE])
    issue_totals[cat] += 1
    if r[C_RESOLVED] == "Yes":
        issue_resolved[cat] += 1

total_resolved = sum(issue_resolved.values())

# ── Print summary ─────────────────────────────────────────────────────────────
print(f"=== Stats Summary ===")
print(f"Responses: {N}")
print(f"Avg all dims: {avg_all}")
print(f"% rated 4–5 overall: {pct_ov}%")
print(f"Faculty avg: {round(mean(faculty),2)}")
print(f"No issues: {pct_no_issues}%")
print(f"Issues reported: {n_issues}, resolved: {total_resolved}")

# ── Patch HTML ────────────────────────────────────────────────────────────────
with open(HTML_PATH, "r", encoding="utf-8") as f:
    html = f.read()

def set_anim_num(html, old_val, old_dec, new_val, new_dec=None, suffix=""):
    """Replace data-val and inner text for an anim-num span."""
    if new_dec is None:
        new_dec = old_dec
    fmt   = f"{new_val:.{new_dec}f}{suffix}"
    pattern = (
        rf'(<span class="anim-num" data-val="{re.escape(str(old_val))}"'
        rf'(?:[^>]*)?>)[^<]*(</span>)'
    )
    new_attrs = f'data-val="{new_val:.{new_dec}f}"'
    if suffix:
        new_attrs += f' data-suffix="{suffix}"'
    if new_dec != old_dec:
        new_attrs += f' data-dec="{new_dec}"'
    return re.sub(
        pattern,
        lambda m: m.group(0).replace(str(old_val), f"{new_val:.{new_dec}f}").replace(m.group(0).split(">")[1].split("<")[0], fmt),
        html, count=1
    )

# ── Build new HTML sections ───────────────────────────────────────────────────

# --- Slide 1 sub text ---
html = re.sub(
    r'\d+ participants responded to the post-event satisfaction survey',
    f'{N} participants responded to the post-event satisfaction survey',
    html, count=1
)

# --- Slide 1 stat cards ---
def replace_anim_num(html, old_val_str, new_val_str, new_text, count=1):
    pattern = (
        rf'data-val="{re.escape(old_val_str)}"([^>]*)>(?:[^<]*)'
    )
    replacement = rf'data-val="{new_val_str}"\1>{new_text}'
    return re.sub(pattern, replacement, html, count=count)

html = replace_anim_num(html, "212", str(N), str(N))  # responses

avg_str = f"{avg_all:.2f}"
html = re.sub(r'data-val="4\.\d+" data-dec="2">4\.\d+(?=</span>.*?out of 5)', f'data-val="{avg_str}" data-dec="2">{avg_str}', html, count=1)

html = re.sub(r'data-val="\d+" data-dec="0" data-suffix="%">\d+%(?=</span>\s*</div>\s*<div class="desc">overall experience)', f'data-val="{pct_ov}" data-dec="0" data-suffix="%">{pct_ov}%', html, count=1)

html = re.sub(r'data-val="\d+" data-dec="0" data-suffix="%">\d+%(?=</span>\s*</div>\s*<div class="desc">of respondents)', f'data-val="{pct_no_issues}" data-dec="0" data-suffix="%">{pct_no_issues}%', html, count=1)

# --- Slide 2: overall rating card ---
def build_star_bars(d, color):
    lines = []
    for star in range(5, 0, -1):
        lines.append(
            f'          <div class="star-row"><span class="lbl">{star}</span>'
            f'<div class="bar-track"><div class="bar-fill" data-w="{d[star]["pct"]}%" '
            f'style="background:{color}"></div></div>'
            f'<span class="cnt">{d[star]["cnt"]}</span></div>'
        )
    return "\n".join(lines)

ov_mean = round(mean(overall), 2)
cm_mean = round(mean(comm_r), 2)
fa_mean = round(mean(faculty), 2)

# Overall card
html = re.sub(
    r'(<div class="dim">Overall experience</div>\s*<div class="mean" style="color:var\(--purple\)"><span class="anim-num" data-val=")[\d.]+(")[^/]*/5</span></div>\s*<div class="pct">\d+% rated 4–5</div>\s*<div class="star-bars">.*?</div>\s*</div>',
    lambda m: (
        f'{m.group(1)}{ov_mean}{m.group(2)}{ov_mean} / 5</span></div>\n'
        f'        <div class="pct">{pct_ov}% rated 4–5</div>\n'
        f'        <div class="star-bars">\n'
        f'{build_star_bars(ov_dist, "var(--purple)")}\n'
        f'        </div>\n'
        f'      </div>'
    ),
    html, count=1, flags=re.DOTALL
)

# Comm card
html = re.sub(
    r'(<div class="dim">Event communication</div>\s*<div class="mean" style="color:var\(--purple-mid\)"><span class="anim-num" data-val=")[\d.]+(")[^/]*/5</span></div>\s*<div class="pct">\d+% rated 4–5</div>\s*<div class="star-bars">.*?</div>\s*</div>',
    lambda m: (
        f'{m.group(1)}{cm_mean}{m.group(2)}{cm_mean} / 5</span></div>\n'
        f'        <div class="pct">{pct_45_comm}% rated 4–5</div>\n'
        f'        <div class="star-bars">\n'
        f'{build_star_bars(cm_dist, "var(--purple-mid)")}\n'
        f'        </div>\n'
        f'      </div>'
    ),
    html, count=1, flags=re.DOTALL
)

# Faculty card
html = re.sub(
    r'(<div class="dim">Faculty advisor support</div>\s*<div class="mean" style="color:var\(--green\)"><span class="anim-num" data-val=")[\d.]+(")[^/]*/5</span></div>\s*<div class="pct">\d+% rated 4–5</div>\s*<div class="star-bars">.*?</div>\s*</div>',
    lambda m: (
        f'{m.group(1)}{fa_mean}{m.group(2)}{fa_mean} / 5</span></div>\n'
        f'        <div class="pct">{pct_45_fac}% rated 4–5</div>\n'
        f'        <div class="star-bars">\n'
        f'{build_star_bars(fa_dist, "var(--green)")}\n'
        f'        </div>\n'
        f'      </div>'
    ),
    html, count=1, flags=re.DOTALL
)

# Faculty callout
html = re.sub(
    r'Faculty advisor satisfaction was the highest-scoring dimension — \d+ of \d+ respondents gave a perfect 5\.',
    f'Faculty advisor satisfaction was the highest-scoring dimension — {fa5} of {fa_n} respondents gave a perfect 5.',
    html
)

# --- Slide 3 sub text ---
html = re.sub(
    r'Bars show share of responses per sentiment category \(n ≈ \d+\)',
    f'Bars show share of responses per sentiment category (n ≈ {N})',
    html
)

# --- Slide 3 tool bars ---
TOOL_LABELS = [name for name in TOOL_NAMES]
GOLD_TOOLS  = {TOOL_NAMES[i] for i in gold_tools}

def build_tool_item(name, t, is_gold):
    cls = "tool-item anim-card gold-strength-row" if is_gold else "tool-item anim-card"
    return (
        f'    <div class="{cls}">\n'
        f'      <div class="tool-lbl">{name}</div>\n'
        f'      <div class="tool-bar">\n'
        f'        <div class="seg" data-w="{t["vh"]}%" style="background:var(--green)"></div>\n'
        f'        <div class="seg" data-w="{t["sh"]}%" style="background:#9FE1CB"></div>\n'
        f'        <div class="seg" data-w="{t["neu"]}%" style="background:#E0DFF5"></div>\n'
        f'        <div class="seg" data-w="{t["prob"]}%"  style="background:var(--red)"></div>\n'
        f'      </div>\n'
        f'      <div class="tool-pct">{t["pos"]}% positive · {t["neu"]}% neutral · {t["prob"]}% problematic</div>\n'
        f'    </div>'
    )

new_tool_html = "\n".join(
    build_tool_item(TOOL_NAMES[i], tools[i], i in gold_tools)
    for i in range(len(TOOL_NAMES))
)

html = re.sub(
    r'(<div class="slide-main">\s*)<div class="tool-item.*?</div>\s*<div class="legend">',
    lambda m: m.group(1) + new_tool_html + '\n\n    <div class="legend">',
    html, count=1, flags=re.DOTALL
)

# --- Slide 5 sub text ---
html = re.sub(
    r'n = \d+, multiple selections allowed',
    f'n = {n_imp}, multiple selections allowed',
    html
)

# --- Slide 5 improvement rows ---
def build_imp_row(label, cnt, n, gold):
    pct_str = round(cnt / n * 100)
    cls = "imp-row anim-card gold-strength-row" if gold else "imp-row anim-card"
    style = "" if gold else ' style="background:var(--purple-mid)"'
    return (
        f'      <div class="{cls}"><span class="imp-lbl">{label}</span>'
        f'<div class="imp-track"><div class="imp-fill" data-w="{pct_str}%"{style}></div></div>'
        f'<span class="imp-val">{cnt} ({pct_str}%)</span></div>'
    )

imp_rows = "\n".join(
    build_imp_row(label, cnt, n_imp, i < 3)
    for i, (label, cnt) in enumerate(improvements)
)
new_imp_html = f'    <div class="imp-list">\n{imp_rows}\n    </div>'

html = re.sub(
    r'<div class="imp-list">.*?</div>\s*</div>(?=\s*</div>\s*<div class="callout)',
    new_imp_html + '\n    </div>',
    html, count=1, flags=re.DOTALL
)

# --- Slide 6 sub text ---
top1 = improvements[0] if improvements else ("", 0)
pct_top1 = round(top1[1] / n_imp * 100) if n_imp else 0
html = re.sub(
    r'\d+ participants \(\d+%\) reported an issue\. The office resolved the majority — \d+ of \d+ confirmed resolved\.',
    f'{n_issues} participants ({round(n_issues/N*100)}%) reported an issue. The office resolved the majority — {total_resolved} of {n_issues} confirmed resolved.',
    html
)

# --- Slide 6 issue cards ---
def res_color(rate):
    if rate >= 75: return "var(--green)"
    if rate >= 50: return "var(--gold)"
    return "var(--red)"

def build_issue_card(cat, t, res, gold):
    rate = round(res / t * 100) if t else 0
    color = res_color(rate)
    cls = "issue-card anim-card gold-strength-row" if gold else "issue-card anim-card"
    plural = "report" if t == 1 else "reports"
    return (
        f'      <div class="{cls}">\n'
        f'        <div style="flex:1"><div class="itype">{cat}</div><div class="icnt">{t} {plural}</div></div>\n'
        f'        <div class="ires" style="color:{color}">{res}/{t} resolved'
        f'<div style="font-size:10px;color:var(--gray);font-weight:400">{rate}% resolution rate</div></div>\n'
        f'        <div class="ibar" style="background:{color}"></div>\n'
        f'      </div>'
    )

issue_cards = "\n".join(
    build_issue_card(
        cat,
        issue_totals[cat],
        issue_resolved[cat],
        cat == "Printing defect"
    )
    for cat in ISSUE_ORDER
    if issue_totals[cat] > 0
)
new_issue_html = f'    <div class="issue-list">\n{issue_cards}\n    </div>'

html = re.sub(
    r'<div class="issue-list">.*?</div>\s*</div>(?=\s*</div>\s*<div class="callout)',
    new_issue_html + '\n    </div>',
    html, count=1, flags=re.DOTALL
)

# ── Write output ──────────────────────────────────────────────────────────────
with open(HTML_PATH, "w", encoding="utf-8") as f:
    f.write(html)

print(f"index.html updated successfully.")
